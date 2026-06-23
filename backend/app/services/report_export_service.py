import io
from datetime import date
from typing import List, Dict, Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
from reportlab.lib.enums import TA_RIGHT, TA_CENTER, TA_LEFT

BRAND_BLUE = colors.HexColor("#1d4ed8")
BRAND_LIGHT = colors.HexColor("#eff6ff")
GREY_TEXT = colors.HexColor("#374151")
GREY_BORDER = colors.HexColor("#e5e7eb")


def _excel_style():
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1d4ed8", end_color="1d4ed8", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="e5e7eb"),
        right=Side(style="thin", color="e5e7eb"),
        top=Side(style="thin", color="e5e7eb"),
        bottom=Side(style="thin", color="e5e7eb"),
    )
    return header_font, header_fill, header_align, thin_border


def generate_excel_report(
    monthly_revenue: List[Dict[str, Any]],
    yearly_revenue: List[Dict[str, Any]],
    invoice_summary: Dict[str, Dict[str, Any]],
    outstanding: List[Dict[str, Any]],
) -> bytes:
    wb = openpyxl.Workbook()
    header_font, header_fill, header_align, thin_border = _excel_style()

    # Sheet 1: Monthly Revenue
    ws1 = wb.active
    ws1.title = "Monthly Revenue"
    headers = ["Month", "Revenue (PKR)", "Invoice Count"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    for row_idx, m in enumerate(monthly_revenue, 2):
        ws1.cell(row=row_idx, column=1, value=m.get("month", "")).border = thin_border
        ws1.cell(row=row_idx, column=2, value=m.get("revenue", 0)).border = thin_border
        ws1.cell(row=row_idx, column=3, value=m.get("invoice_count", 0)).border = thin_border
    ws1.column_dimensions["A"].width = 20
    ws1.column_dimensions["B"].width = 25
    ws1.column_dimensions["C"].width = 18

    # Sheet 2: Yearly Revenue
    ws2 = wb.create_sheet("Yearly Revenue")
    headers = ["Year", "Revenue (PKR)", "Invoice Count"]
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    for row_idx, y in enumerate(yearly_revenue, 2):
        ws2.cell(row=row_idx, column=1, value=y.get("year", "")).border = thin_border
        ws2.cell(row=row_idx, column=2, value=y.get("revenue", 0)).border = thin_border
        ws2.cell(row=row_idx, column=3, value=y.get("invoice_count", 0)).border = thin_border
    ws2.column_dimensions["A"].width = 15
    ws2.column_dimensions["B"].width = 25
    ws2.column_dimensions["C"].width = 18

    # Sheet 3: Invoice Summary
    ws3 = wb.create_sheet("Invoice Summary")
    headers = ["Status", "Count", "Amount (PKR)"]
    for col, h in enumerate(headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    for row_idx, (status, data) in enumerate(invoice_summary.items(), 2):
        ws3.cell(row=row_idx, column=1, value=status.replace("_", " ").title()).border = thin_border
        ws3.cell(row=row_idx, column=2, value=data.get("count", 0)).border = thin_border
        ws3.cell(row=row_idx, column=3, value=data.get("amount", 0)).border = thin_border
    ws3.column_dimensions["A"].width = 25
    ws3.column_dimensions["B"].width = 15
    ws3.column_dimensions["C"].width = 25

    # Sheet 4: Outstanding
    ws4 = wb.create_sheet("Outstanding")
    headers = ["Client Name", "Total Invoiced", "Total Paid", "Outstanding"]
    for col, h in enumerate(headers, 1):
        cell = ws4.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    for row_idx, o in enumerate(outstanding, 2):
        ws4.cell(row=row_idx, column=1, value=o.get("client_name", "")).border = thin_border
        ws4.cell(row=row_idx, column=2, value=o.get("total_invoiced", 0)).border = thin_border
        ws4.cell(row=row_idx, column=3, value=o.get("total_paid", 0)).border = thin_border
        ws4.cell(row=row_idx, column=4, value=o.get("outstanding", 0)).border = thin_border
    ws4.column_dimensions["A"].width = 30
    ws4.column_dimensions["B"].width = 22
    ws4.column_dimensions["C"].width = 22
    ws4.column_dimensions["D"].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _pdf_styles():
    base = getSampleStyleSheet()
    return {
        "title": ParagraphStyle("title", fontSize=18, textColor=BRAND_BLUE,
                                fontName="Helvetica-Bold", spaceAfter=16),
        "h2": ParagraphStyle("h2", fontSize=12, textColor=BRAND_BLUE,
                             fontName="Helvetica-Bold", spaceBefore=12, spaceAfter=6),
        "body": ParagraphStyle("body", fontSize=9, textColor=GREY_TEXT,
                               fontName="Helvetica", leading=14),
        "small": ParagraphStyle("small", fontSize=8, textColor=GREY_TEXT,
                                fontName="Helvetica", leading=12),
        "footer": ParagraphStyle("footer", fontSize=8, textColor=colors.HexColor("#9ca3af"),
                                 fontName="Helvetica", alignment=TA_CENTER),
    }


def generate_pdf_report(
    monthly_revenue: List[Dict[str, Any]],
    yearly_revenue: List[Dict[str, Any]],
    invoice_summary: Dict[str, Dict[str, Any]],
    outstanding: List[Dict[str, Any]],
) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=18*mm, rightMargin=18*mm,
                            topMargin=14*mm, bottomMargin=14*mm)
    S = _pdf_styles()
    W = A4[0] - 36*mm
    story = []

    story.append(Paragraph("Financial Report", S["title"]))
    story.append(Paragraph(f"Generated: {date.today().strftime('%d/%m/%Y')}", S["body"]))
    story.append(Spacer(1, 8))
    story.append(HRFlowable(width="100%", thickness=2, color=BRAND_BLUE, spaceAfter=12))

    # Monthly Revenue
    story.append(Paragraph("Monthly Revenue", S["h2"]))
    if monthly_revenue:
        col_widths = [W*0.4, W*0.35, W*0.25]
        rows = [
            [Paragraph("Month", S["small"]), Paragraph("Revenue (PKR)", S["small"]),
             Paragraph("Count", S["small"])],
        ]
        for m in monthly_revenue:
            rows.append([
                Paragraph(m.get("month", ""), S["small"]),
                Paragraph(f"{m.get('revenue', 0):,.0f}", S["small"]),
                Paragraph(str(m.get("invoice_count", 0)), S["small"]),
            ])
        tbl = Table(rows, colWidths=col_widths, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), BRAND_BLUE),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.3, GREY_BORDER),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ]))
        story.append(tbl)
    else:
        story.append(Paragraph("No data available.", S["body"]))

    story.append(Spacer(1, 10))

    # Yearly Revenue
    story.append(Paragraph("Yearly Revenue", S["h2"]))
    if yearly_revenue:
        col_widths = [W*0.4, W*0.35, W*0.25]
        rows = [
            [Paragraph("Year", S["small"]), Paragraph("Revenue (PKR)", S["small"]),
             Paragraph("Count", S["small"])],
        ]
        for y in yearly_revenue:
            rows.append([
                Paragraph(str(y.get("year", "")), S["small"]),
                Paragraph(f"{y.get('revenue', 0):,.0f}", S["small"]),
                Paragraph(str(y.get("invoice_count", 0)), S["small"]),
            ])
        tbl = Table(rows, colWidths=col_widths, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), BRAND_BLUE),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.3, GREY_BORDER),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ]))
        story.append(tbl)
    else:
        story.append(Paragraph("No data available.", S["body"]))

    story.append(Spacer(1, 10))

    # Invoice Summary
    story.append(Paragraph("Invoice Status Breakdown", S["h2"]))
    if invoice_summary:
        col_widths = [W*0.4, W*0.25, W*0.35]
        rows = [
            [Paragraph("Status", S["small"]), Paragraph("Count", S["small"]),
             Paragraph("Amount (PKR)", S["small"])],
        ]
        for status, data in invoice_summary.items():
            rows.append([
                Paragraph(status.replace("_", " ").title(), S["small"]),
                Paragraph(str(data.get("count", 0)), S["small"]),
                Paragraph(f"{data.get('amount', 0):,.0f}", S["small"]),
            ])
        tbl = Table(rows, colWidths=col_widths, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), BRAND_BLUE),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.3, GREY_BORDER),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ]))
        story.append(tbl)
    else:
        story.append(Paragraph("No data available.", S["body"]))

    story.append(Spacer(1, 10))

    # Outstanding
    story.append(Paragraph("Outstanding by Client", S["h2"]))
    if outstanding:
        col_widths = [W*0.35, W*0.22, W*0.2, W*0.23]
        rows = [
            [Paragraph("Client", S["small"]), Paragraph("Invoiced", S["small"]),
             Paragraph("Paid", S["small"]), Paragraph("Outstanding", S["small"])],
        ]
        for o in outstanding:
            rows.append([
                Paragraph(o.get("client_name", ""), S["small"]),
                Paragraph(f"{o.get('total_invoiced', 0):,.0f}", S["small"]),
                Paragraph(f"{o.get('total_paid', 0):,.0f}", S["small"]),
                Paragraph(f"{o.get('outstanding', 0):,.0f}", S["small"]),
            ])
        tbl = Table(rows, colWidths=col_widths, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), BRAND_BLUE),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.3, GREY_BORDER),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ]))
        story.append(tbl)
    else:
        story.append(Paragraph("No data available.", S["body"]))

    story.append(Spacer(1, 20))
    story.append(HRFlowable(width="100%", thickness=0.5, color=GREY_BORDER, spaceAfter=4))
    story.append(Paragraph(
        "Generated by Crop2X CRM  |  Crop2X (Private) Limited",
        S["footer"],
    ))

    doc.build(story)
    return buf.getvalue()
